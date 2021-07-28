
'''
<계기>
현재 실험실 내 Chemicals List는 Cas No가 기재되어 있지 않다.
화학물질은 학명이 다양하기 때문에 Cas No가 없으면 동일 물질임을 한 눈에 확인하기 어렵다.
실험실 내 화학물질 관리를 위해 Cas No 추가는 필수적이지만,
손으로 하나하나 학명을 검색하여 추가하는 것은 매우 비효율적이다.

그에 따라 자동적으로 엑셀파일에서 학명을 검색하여 Cas No를 찾고
그 결과를 엑셀파일의 해당되는 셀에 입력하는 코드를 작성하기로 했다.

<프로세스>
openpyxl로 엑셀파일을 열어 학명을 얻어오고
selenium과 chromedriver를 사용하여 Cas No를 찾는다.

'''

# 필요한 라이브러리를 불러온다.
import pandas as pd
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
import time


# 열고자하는 역셀 파일 경로를 설정한다.
location = 'C:/Users/GhostHowling/Downloads'
file = '202103-Updated Inventory Chemicals List.xlsx'

# 지정한 파일을 pandas로 연 뒤 데이터를 변수에 지정한다.
data_pd = pd.read_excel('{}/{}'.format(location,file),
              header=None, index_col=None, names=None)
data_np = pd.DataFrame.to_numpy(data_pd)

# 엑셀파일을 연 뒤, 각 시트를 불러와서 변수에 지정한다.
wb = openpyxl.load_workbook('{}/{}'.format(location,file))
ws1 = wb['Inside Lab']
ws2 = wb['Outside Lab']
ws3 = wb['2021년 ']

# 각 시트의 마지막 행이 어디까지인지 확인한다.
# print(ws1.max_row)
# print(ws2.max_row)
# print(ws3.max_row)

# 1번째 줄부터 반복문을 수행하도록 한다.
cnt = 1
flag = True
while(flag):
    # A컬럼과 B컬럼의 값을 변수 num, des에 지정한다.
    num = ws3.cell(row=cnt, column=1).value
    des = ws3.cell(row=cnt, column=2).value
    # num이 숫자이고, des가 공백이 아닌 문자일 때만 출력, 검색한다.
    if type(des) is str:
        if type(num) is int:
            print('No {}: {}'.format(num, des))

            # chromedriver를 변수에 지정한다.
            driver = webdriver.Chrome('./webdriver/chromedriver.exe')
            driver.implicitly_wait(2)

            # google 페이지에 접근하여 화합물명에 해당하는 cas no를 검색한다.
            driver.get('https://google.com')
            q = driver.find_element_by_name('q')
            q.send_keys(des+' cas number')
            q.submit()

            # 검색 이후의 웹 페이지 url을 변수에 지정한 뒤, driver로 html 페이지를 가져온다.
            url = driver.current_url
            driver.get(url)
            html = driver.page_source

            # if '이것을 찾으셨나요' in html:
            #     a = driver.find_element_by_xpath("//a[@href=")
            #     driver.get(a[0].get_attribute('href'))
            #     print('클릭 필요') # p.gqLncc.card-section.KDCVqf a.gL9Hy
            

            # 파싱
            soup = BeautifulSoup(html, 'html.parser')
            try:
                anwser = soup.select_one('div#search div.IZ6rdc')
                print(anwser.text)
            except:
                try:
                    anwser = soup.select_one('div#search div.Z0LcW.XcVN5d')
                    print(anwser.text)
                except:
                    print('다른 방법 추구')



    cnt += 1
    # 시트의 마지막 줄을 넘어가면 종료한다.
    if cnt>ws3.max_row:
        break



print('종료')
